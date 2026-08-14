"""
openapi_to_csv.py — แปลง openapi.json เป็น CSV และ Excel (.xlsx)

วิธีใช้:
    # ต้องรัน FastAPI server ก่อน แล้วค่อยรัน script นี้
    python scripts/openapi_to_csv.py

    # หรือระบุ URL เอง
    python scripts/openapi_to_csv.py --url http://localhost:8000/openapi.json

ผลลัพธ์:
    scripts/api_snapshot.csv
    scripts/api_snapshot.xlsx
"""

import argparse
import csv
import json
import urllib.request
from pathlib import Path

COLUMNS = ["Method", "Path", "Summary", "Description", "Tags", "Auth Required"]
DEFAULT_URL = "http://localhost:8000/openapi.json"


def fetch_openapi(url: str) -> dict:
    """ดาวน์โหลด openapi.json จาก server."""
    with urllib.request.urlopen(url) as response:
        return json.loads(response.read().decode())


def parse_endpoints(spec: dict) -> list[dict]:
    """แปลง OpenAPI spec เป็น list ของ endpoints."""
    rows: list[dict] = []
    paths = spec.get("paths", {})

    for path, methods in paths.items():
        for method, detail in methods.items():
            if method in ("get", "post", "put", "patch", "delete"):
                # ตรวจว่าต้อง auth หรือไม่ จาก security field
                has_auth = bool(detail.get("security"))
                rows.append({
                    "Method": method.upper(),
                    "Path": path,
                    "Summary": detail.get("summary", ""),
                    "Description": detail.get("description", ""),
                    "Tags": ", ".join(detail.get("tags", [])),
                    "Auth Required": "Yes" if has_auth else "No",
                })

    return rows


def write_csv(rows: list[dict], output_path: Path) -> None:
    """เขียน CSV ไฟล์."""
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"✅ CSV saved: {output_path}")


def write_excel(rows: list[dict], output_path: Path) -> None:
    """เขียน Excel (.xlsx) ไฟล์."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("⚠️  openpyxl not installed — skipping Excel export")
        print("   Install with: uv add openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "API Snapshot"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write header
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(COLUMNS, 1):
        max_len = len(col_name)
        for row in rows:
            val = str(row.get(col_name, ""))
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 60)

    wb.save(output_path)
    print(f"✅ Excel saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Export OpenAPI spec to CSV & Excel")
    parser.add_argument("--url", default=DEFAULT_URL, help=f"OpenAPI JSON URL (default: {DEFAULT_URL})")
    args = parser.parse_args()

    script_dir = Path(__file__).parent

    print(f"📡 Fetching OpenAPI spec from {args.url} ...")
    spec = fetch_openapi(args.url)
    print(f"📋 API title: {spec.get('info', {}).get('title', 'N/A')}")

    rows = parse_endpoints(spec)
    print(f"📊 Found {len(rows)} endpoints")

    write_csv(rows, script_dir / "api_snapshot.csv")
    write_excel(rows, script_dir / "api_snapshot.xlsx")

    print("🎉 Done!")


if __name__ == "__main__":
    main()
